import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1'] # return first sheet
    #cell = sheet['a1']   # first cell coordinates a1
    #cell = sheet.cell(1,1)   # (*,*) pass the row and column
    #print(cell.value) 
    #print(sheet.max_row)
    #print(sheet.max_column)

    for row in range(2, sheet.max_row + 1):  # ignore the first row 
        cell = sheet.cell(row, 3)
        corr_price = cell.value * 0.9
        corr_price_cell = sheet.cell(row, 4)
        corr_price_cell.value = corr_price

    values= Reference(sheet,     # select values to add into the chart
            min_row = 2, 
            max_row = sheet.max_row,
            min_col =4,
            max_col =4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)


process_workbook('transactions.xlsx')